from influxdb import InfluxDBClient
from influxdb.client import InfluxDBClientError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,numbers
from datetime import datetime

DATE = datetime.now().strftime("%m-%Y")
SERIES = ""

def initialization_influx():
    INFLUX_HOST = 'localhost'
    # INFLUX_HOST = 't13363.astra-honda.com'  
    INFLUX_HOST = '10.9.12.38'        
    INFLUX_PORT = '8086'
    INFLUX_USERNAME = 'admin'
    INFLUX_PASSWORD = 'Honda2020!'
    INFLUX_PASSWORD = 'ITIOT2019!'    
    INFLUX_DB = 'AHMITIOT'

    global SERIES
    try:
        SERIES = InfluxDBClient(INFLUX_HOST, INFLUX_PORT)
        SERIES = InfluxDBClient(INFLUX_HOST, INFLUX_PORT, INFLUX_USERNAME,
                                INFLUX_PASSWORD)
            
        SERIES.switch_database(INFLUX_DB)
        return True
    except Exception:
        print("Cannot connect influxDB!")
        return False


def select_series(query):
    try:
        RESULT = SERIES.query(query)
        return RESULT
    except Exception as e:
        print("An exception occurred ::", e)
        return False

def service_maintenance():        
    QUERY = f'SELECT VASSETID,VDESC,NAVGPERDAY,NREMRUNHOUR,DNEXTMTC,DMODI FROM "AHMITIOT_DTLSCHMTCS" '
    results = select_series(QUERY)     
    # Create Excel Report
    ROWEXCEL = 1
    workbook = Workbook()
    sheet = workbook.active 
    sheet.title = "Data Report"
    sheet.cell(row=ROWEXCEL, column=1).value = "Asset ID"
    sheet.cell(row=ROWEXCEL, column=2).value = "Asset Desc"
    sheet.cell(row=ROWEXCEL, column=3).value = "Runhour Harian"
    sheet.cell(row=ROWEXCEL, column=4).value = "Runhour Tersisa"
    sheet.cell(row=ROWEXCEL, column=5).value = "Est Next Maintenance"
    sheet.cell(row=ROWEXCEL, column=1).font = Font(bold=True)
    sheet.cell(row=ROWEXCEL, column=2).font = Font(bold=True)
    sheet.cell(row=ROWEXCEL, column=3).font = Font(bold=True)
    sheet.cell(row=ROWEXCEL, column=4).font = Font(bold=True)
    sheet.cell(row=ROWEXCEL, column=5).font = Font(bold=True)        

    for measurement in results.get_points():
        ROWEXCEL+=1
        VASSETID = measurement['VASSETID'] 
        VDESC = measurement['VDESC']                                                   
        NAVGPERDAY = measurement['NAVGPERDAY']
        NREMRUNHOUR = measurement['NREMRUNHOUR']
        DNEXTMTC = measurement['DNEXTMTC']        

        # date_DMODI = datetime.strptime(DMODI, '%d/%m/%Y %H:%M:%S')        
        date_DNEXTMTC = datetime.strptime(DNEXTMTC, '%d/%m/%Y').date()
        print (f"{VASSETID} | {VDESC} | {NAVGPERDAY:.2f} | {NREMRUNHOUR:.2f} | {date_DNEXTMTC}") 
        sheet.cell(row=ROWEXCEL, column=1).font = Font(bold=False)
        sheet.cell(row=ROWEXCEL, column=2).font = Font(bold=False)
        sheet.cell(row=ROWEXCEL, column=3).font = Font(bold=False)
        sheet.cell(row=ROWEXCEL, column=4).font = Font(bold=False)   
        sheet.cell(row=ROWEXCEL, column=5).font = Font(bold=False)             
        sheet.cell(row=ROWEXCEL, column=1).value = VASSETID        
        sheet.cell(row=ROWEXCEL, column=2).value = VDESC              
        sheet.cell(row=ROWEXCEL, column=3).value = round(NAVGPERDAY,2)
        sheet.cell(row=ROWEXCEL, column=4).value = round(NREMRUNHOUR,2)        
        sheet.cell(row=ROWEXCEL, column=5).value = date_DNEXTMTC     
        sheet.cell(row=ROWEXCEL, column=5).number_format = numbers.FORMAT_DATE_XLSX15           

    sheet.column_dimensions[get_column_letter(1)].width = 21    
    sheet.column_dimensions[get_column_letter(2)].width = 40
    sheet.column_dimensions[get_column_letter(3)].width = 20
    sheet.column_dimensions[get_column_letter(4)].width = 15
    sheet.column_dimensions[get_column_letter(5)].width = 20
    sheet.auto_filter.ref = sheet.dimensions
    # workbook.save(filename=f"list_next_maintenance_AC-{DATE}.xlsx")
    workbook.save(filename=f"/home/nifi/ahmitiot011/list_next_maintenance_AC-{DATE}.xlsx")

def main():
    initialization_influx()
    service_maintenance()
    

if __name__ == '__main__':
    main()